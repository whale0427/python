import openpyxl,datetime

class ReadExcel:
    def __init__(self,file_path):
        self.workbook=openpyxl.load_workbook(file_path)
        self.worksheet=self.workbook.active

    def get_data(self):
        datas=[]
        for row in self.worksheet.iter_rows():
            data_row=[]
            for cell in row:
                data_row.append(cell.value)
            datas.append(data_row)
        return datas

class WriteExcel:
    def __init__(self,file_path,datas):
        self.file_path = file_path
        self.datas = datas
        self.workbook=openpyxl.Workbook()
        self.worksheet=self.workbook.active

    def add_excel(self):
        for row in self.datas:
            self.worksheet.append(row)
        self.workbook.save(self.file_path)
